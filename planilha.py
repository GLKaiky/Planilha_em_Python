import atexit
import os
import openpyxl
import	requests
import shutil
from openpyxl.styles import NamedStyle, Font, Alignment, numbers, PatternFill, Border, Side
from PIL import Image
from datetime import datetime

#Funções definidas ao codigo

def baixar_arquivo(url):
    nome_arquivo = 'DemonstrativoFinanceiro.xlsx'  # Obtém o nome do arquivo a partir da URL
    caminho_completo = f'{r"/home/glkaiky/Desktop/Planilha_em_Python"}/{nome_arquivo}'  # Caminho completo do arquivo de destino

    # Faz o download do arquivo
    r = requests.get(url, stream=True)
    if r.status_code == 200:
        with open(caminho_completo, 'wb') as f:
            r.raw.decode_content = True
            shutil.copyfileobj(r.raw, f)
        print(f'O arquivo foi baixado para {caminho_completo}')
    else:
        print('Falha ao baixar o arquivo.')

def Tipo_Moeda(celula, valor):

    fonte = Font(name='Calibri', bold=True, color='000000')
    alinhamento = Alignment(horizontal='right', vertical='center')
    preenchimento = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    celula.font = fonte
    celula.alignment = alinhamento
    celula.fill = preenchimento
    celula.number_format = '"R$"#,##0.00_);[Red]("$"#,##0.00)'
    celula.value = valor

def Bordas(bloco):
    célula = sheet[bloco]
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))    
    célula.border = borda

def receitas():

    Total = 0
    planilha = ['B5', 'B6', 'B7', 'B8', 'B9', 'B10']
    planilha2 = ['C5', 'C6', 'C7', 'C8', 'C9', 'C10']
    apto = [101, 102, 201, 202, 301, 302]
    Condominio = input('Defina o valor:\n')
    Condominio = Condominio.replace(',', '.')
    Numerico = float(Condominio)
    Condominio_Formatado = f'R${Numerico:.2f}'

    for cell in planilha + planilha2:
        Background_Color(cell, 'ffffff')
        Bordas(cell)
        cell = sheet[cell]
        cell.value = ''

    for i, condomino in enumerate(apto):
        while True:
            Pagou = input(f'Apartamento {condomino} pagou?:\n')
            Pagou = Pagou.upper()
            if (Pagou == 'SIM'):
               # Formatação_de_Texto(Css, 'Calibri', 16, False, planilha[i], Condominio_Formatado)
                Background_Color(planilha[i], '28eb13')
                Bordas(planilha[i])
                Total+=int(Condominio)
                cell = sheet[planilha[i]]
                cell.value = Condominio_Formatado
                print(f'Pagamento do {condomino} Registrado com sucesso')      
                break
            elif(Pagou == 'NAO' or Pagou == 'NÃO'):
                #Formatação_de_Texto(Css, 'Calibri', 16, False, planilha2[i], 'Não pagou')
                Background_Color(planilha2[i], 'f20c27')
                Bordas(planilha2[i])
                cell = sheet[planilha2[i]]
                cell.value = 'Não Pago'
                print(f'Caloteiro do {condomino} registrado com sucesso')
                break
            else:
                print('Porfavor digite sim ou nao')
    Local = f'{chr(64+2)}{12+8}'
    return Total, Local

def saldo(bloco):
    #Dar acesso a p lanilha

    #Verificar a existencia do estilo "Css"
    Css = None
    Saldo_do_Mes = input('Digite o saldo do Mês anterior:\n')
    print('Saldo definido para: ', 'R$' + Saldo_do_Mes + ',00')
    Saldo_do_Mes = Saldo_do_Mes.replace(',', '.')
    Numerico = float(Saldo_do_Mes)
    Saldo_Formatado = f'R${Numerico:.2f}'

    #Definir localização
    #Formatação_de_Texto(Css, 'Calibri', 20, False, bloco, Saldo_Formatado)

    cell = sheet[bloco]
    cell.value = Saldo_Formatado
    Background_Color('B2', 'eaf51b')
    Bordas(bloco)
    Local = f'{chr(64+2)}{2+17}'
    return Saldo_Formatado, Local

def Background_Color(bloco, cor):

    #Acesso a planilha
    sheet = workbook.active
    fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
    
    #Onde será aplicado o backgroundColor
    cell = sheet[bloco]
    cell.fill = fill

def Definir_Data(bloco):

    if bloco != ' ':
        Background_Color(bloco, 'ffffff')
        Bordas(bloco)
        cell = sheet[bloco]
        cell.value = ''

    # Obter a data atual
    data_atual = datetime.now().date()
    data_atual_formatada = data_atual.strftime('%m-%Y')

    #Formatação_de_Texto(Css, 'Calibri', 16, False, bloco, data_atual_formatada)
    cell = sheet[bloco]
    cell.value = data_atual_formatada
     # Aplicar o estilo de fundo
    Background_Color("A3", "95b3d7")

def open_theend():
    os.system(f'xdg-open "{"/home/glkaiky/Desktop/Planilha_em_Python/DemonstrativoFinanceiro.xlsx"}"')

def Despesas():

    Marcador = 14
    Soma = 0
    Local2 = 0
    #sheet.delete_rows(16, 6) #(Deletar linhas, ainda nao achei solução para fazer isso sozinho)
    Adicionar = input('Adicionar despesas?')
    Adicionar = Adicionar.upper()

    while Adicionar == 'SIM' or Adicionar == 'S':
        if(Adicionar == 'SIM' or Adicionar == 'S'):
            Marcador+=1
            sheet.insert_rows(Marcador)
            Despesa = input('Digite a despesa:\n')
            Valor = input('Digite o valor:\n')
            
            Valor = Valor.replace(',', '.')
            Numerico = float(Valor)
            Soma+=Numerico
            Valor_Formatado = f'R${Numerico:.2f}'
            cell = sheet.cell(row=Marcador, column=1)
            cell2 = sheet.cell(row=Marcador, column=2)  
            Local = f'{chr(64+2)}{Marcador}'
            Local2 = f'{chr(64+2)}{Marcador + 2}'
            Local3 = f'{chr(64+1)}{Marcador}'
            Background_Color(Local, 'f20c27')
            Bordas(Local)
            Bordas(Local3)
            cell2.value = Valor_Formatado
            cell.value = Despesa
    
            
        Adicionar = input('Adicionar mais despesas? (SIM ou NÃO)').upper()
    return Soma, Local2

try:
    #baixar arquivo
    baixar_arquivo('https://docs.google.com/uc?id=1CiJb0QKFnC4qq5Cc9l7Xj1GBNfTZDVHX&export=download')
    # Abrir planilha para edição
    workbook = openpyxl.load_workbook('DemonstrativoFinanceiro.xlsx')
    workbook.create_sheet(title='ConfirmaçõesPagamentos')
    print('Arquivo aberto com sucesso, edite como quiser!')


    # Acessando planilha ativa
    sheet = workbook.active


    #Definir o saldo do Mês
    S2, Local = saldo("B2")
    cell = sheet[Local]
    cell.value = S2

    #Adicionar data do dia ao Bloco
    Definir_Data("A3")
   
    Valor, Local = receitas()
    print(Valor)
    Numerico = float(Valor)
    Valor_Formatado = f'R${Numerico:.2f}'
    cell = sheet['B12']
    cell.value = Valor_Formatado
    cell = sheet[Local]
    cell.value = Valor_Formatado
    Receitas_total = Valor_Formatado

    Resultado, Local= Despesas()
    print(Resultado)
    print(Local)
    Numerico = float(Resultado)
    Resultado_Formatado = f'R${Numerico:.2f}'
    cell = sheet[Local]
    cell.value = Resultado_Formatado
    sheet.cell(row=sheet.max_row, column=2, value=Resultado_Formatado)

    Final_Line = sheet.max_row
    sheet.insert_rows(Final_Line + 2)

    cell = sheet.cell(row=Final_Line + 2, column=1, value = 'Saldo Atual:')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    fonte_estilizada = Font(name='Calibri', size=23, bold=False, color='000000')
    cell.font = fonte_estilizada
    Bordas(f"B{Final_Line+2}")
    Background_Color(f"A{Final_Line+2}", '9bbb59')

    S2 = float(''.join(filter(str.isdigit, S2)))
    Receitas_total = float(''.join(filter(str.isdigit,Receitas_total)))
    Resultado = float(''.join(filter(str.isdigit,Resultado_Formatado)))

    # Realize o cálculo do "Saldo Atual"
    Saldo_Atual = S2 + Receitas_total - Resultado

    # Formate o resultado
    Saldo_Atual/=100

    Saldo_Atual = round(Saldo_Atual, 2)
    Saldo_Atual_Formatado = f'R${Saldo_Atual}'

    # Insira o resultado na célula D2
    cell = sheet.cell(row=Final_Line + 2, column=2, value = Saldo_Atual_Formatado)
    cell.value = Saldo_Atual_Formatado
    fonte_estilizada = Font(name='Calibri', size=22, bold=False, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = fonte_estilizada
    Background_Color(f"B{Final_Line+2}", 'eaf51b')
    Bordas(f"B{Final_Line+2}")
    
    # Salvar as alterações no arquivo
    workbook.save('DemonstrativoFinanceiro.xlsx')
    
    print('Alterações salvas com sucesso!')

    workbook.close()
    atexit.register(open_theend)

except FileNotFoundError:
    print('Arquivo não encontrado, verificar o erro')
except openpyxl.utils.exceptions.InvalidFileException:
    print('Formato não suportado, converta para xlsx')
